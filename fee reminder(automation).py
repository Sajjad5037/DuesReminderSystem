import openpyxl
import smtplib
import sys
import os

#open the excel sheet
filePath=r"E:\Python\Python Projects\Chapter 16(sending email and texts)\dueRecords.xlsx"
wb=openpyxl.load_workbook(filePath)
sheet=wb.active
lastCol=sheet.max_column
latestMonth=sheet.cell(row=1,column=lastCol).value

#find all unpaid members
UnpaidMembers={}
all_members={}
for r in range(2,sheet.max_row+1): #checking for name and email
    name=sheet.cell(row=r,column=1).value
    email=sheet.cell(row=r,column=2).value
    all_members[name]=email
    unpaid=False
    #check each month's payment status
    for c in range(3,sheet.max_column+1): #assuming month column's start from 3 (lookinf for "paid" "unpaid")
        payment=sheet.cell(row=r,column=c).value
        if payment!='Paid':
            unpaid=True
            break
    if unpaid:
        UnpaidMembers[name]=email
#send customized email reminders
smtpObj=smtplib.SMTP('smtp.gmail.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("proactive1.san@gmail.com","vsjv dmem twvz avhf")

for name,email in UnpaidMembers.items():
    body=f"Subject:Dues Unpaid.\n\nRecords show that you have not paid your fees for one or more months.kindly clear your dues"
    print(f"sending email to {email}")
    sendMailStatus=smtpObj.sendmail("proactive1.san@gmail.com",email,body)
    if sendMailStatus!={}:
        print(f"there was a problem sending email to {email}:{sendMailStatus}")
smtpObj.quit()
